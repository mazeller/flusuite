#!/usr/bin/python

#for Python3
#DEPENDS ON OPENPYXL!
import sys
import getopt
import textwrap

import dateutil
import dateutil.parser as dateParser

import openpyxl
from openpyxl import Workbook
from openpyxl import formatting
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import (
	get_column_letter
)

#HELP DOCUMENATION
def displayHelp():
		print("""
@@ snake @@  11/96
                          _,..,,,_ 
                     '``````^~"-,_`"-,_
       .-~c~-.                    `~:. ^-.     
   `~~~-.c    ;                      `:.  `-,     _.-~~^^~:.
         `.   ;      _,--~~~~-._       `:.   ~. .~          `.
          .` ;'   .:`           `:       `:.   `    _.:-,.    `.
        .' .:   :'    _.-~^~-.    `.       `..'   .:      `.    '
       :  .' _:'   .-'        `.    :.     .:   .'`.        :    ;
  jgs  :  `-'   .:'             `.    `^~~^`   .:.  `.      ;    ;
        `-.__,-~                  ~-.        ,' ':    '.__.`    :'
                                     ~--..--'     ':.         .:'
                                                     ':..___.:'
from: http://chris.com/ascii/index.php?art=animals/reptiles/snakes
		
USAGE: TBD
""")

#Predefine all our colors in one place
clusters = {
    'turkey':['NHNNYR'],
    'purple':['KHHNYR','KHKNYS'],
    'gray':['KTHKYS','NTHNFK','NTQKFN','NTHKFN'],
    'darkgray':['KTHNFK','KTHNSK'],
    'lightpink':['KHQKYR'],
    'gold':['KYNNNK'],
    'red':['NYNNYK','NYNNHK','NHNNYK','NYHNYK'],
    'darkpink':['KHKNYR'],
    'cyan':['NNNDYR', 'NHSNYR', 'NHNNYR', 'NHNDYR'],
    'brown':['NYHGHE'],
    'darkgreen':['KYKNYE'],
    'lightgreen':['KYNNYK'],
    'orange':['KHKNYE'],
    'blue':['KHNNYK'],
    'lightblue':['KHKEYS','KHHNYR']
}

conditionalFill = {
    'turkey': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
    'purple': PatternFill(start_color='800080', end_color='800080', fill_type='solid'),
    'gray': PatternFill(start_color='cccccc', end_color='cccccc', fill_type='solid'),
    'darkgray': PatternFill(start_color='888888', end_color='888888', fill_type='solid'),
    'lightpink': PatternFill(start_color='ffc0cb', end_color='ffc0cb', fill_type='solid'),
    'gold': PatternFill(start_color='f9d624', end_color='f9d624', fill_type='solid'),
    'red': PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'),
    'darkpink': PatternFill(start_color='880088', end_color='880088', fill_type='solid'),
    'cyan':PatternFill(start_color='00ffff', end_color='00ffff', fill_type='solid'),
    'brown':PatternFill(start_color='f4a460', end_color='f4a460', fill_type='solid'),
    'darkgreen':PatternFill(start_color='008800', end_color='008800', fill_type='solid'),
    'lightgreen':PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
    'orange': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
    'blue': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),
    'lightblue': PatternFill(start_color='6495ed', end_color='6495ed', fill_type='solid')
}

#Condtional Formattings for processed page
def setProcessedConditionalFormatting(wsProc):

	for key,list in clusters.items():
		for clusterMotif in list:
				dxf = DifferentialStyle(fill=conditionalFill[key])
				rule = Rule(type="containsText", operator="containsText", text=clusterMotif, dxf=dxf) #blue
				rule.formula = ['NOT(ISERROR(SEARCH("' + clusterMotif + '",D1)))']
				wsProc.conditional_formatting.add('A:A', rule)

#Convert sequence into a color
def assignColor(motif):
	antigenicColor = "unknown"
	
	for key,list in clusters.items():
		for clusterMotif in list:
			if(motif == clusterMotif):
				antigenicColor = key;
				return antigenicColor;

	return antigenicColor
	
def main():
	#Get arguments
	argv = sys.argv[1:]
	yearMotifs = {}
	heuristicAlign = True
	inputfile = ''		#Input .tsv as STRING
	
	try:
		opts, args = getopt.getopt(argv,"i:nh",["input=","noalign","help"])
	except getopt.GetoptError:
		displayHelp()
		sys.exit(2)
	for opt, arg in opts:
		if opt in ('-h',"help"):
			displayHelp()
			sys.exit()
		elif opt in ("-i", "--ifile"):
			inputfile = arg
		elif opt in ("-n", "--noalign"):
			heuristicAlign = False

	#Load input for line by line reading
	with open(inputfile) as f:
		for index, line in enumerate(f):
			col = []
			col = line.split("\t")
			
			#Get Indices from first line
			if(index ==0):
				tsvAccession = col.index("Sequence Accession")
				tsvDate = col.index("Collection Date")
				tsvName = col.index("Strain Name")
				tsvSeq = col.index("Sequence")
					
				#Next iteration
				continue

			#ALIGN -- Use heuristics to 'align' sequences on the fly. Based on calculating offset by identifying the most stable part(s) of the sequence, then shifts left. Will fail on indel
			if(heuristicAlign == True):
				if(col[tsvSeq][0] != "M"):	#96% start properly
					seqShift = col[tsvSeq].upper().find('FSRLNWL')			#Find the relative position of 148(+16)-154(+16) and shift right with spaces
					if(seqShift < (148+16)):
						charFill = ' '*((148+16) - seqShift - 1)
						col[tsvSeq] = charFill + col[tsvSeq]
					elif(seqShift == -1):									 #can either try another relative position or scrap the sequence. I am lazy this time, so I am down with scrapping. Need to insert blanks to keep sheet functioning on the index
						wsOrig.append([])
						wsProc.append([])
						continue

			#Generate motif to figure out color
			try:
				motif = col[tsvSeq][145+16-1] + col[tsvSeq][155+16-1] + col[tsvSeq][156+16-1] + col[tsvSeq][158+16-1] + col[tsvSeq][159+16-1] + col[tsvSeq][189+16-1]
			except:
				motif = ''
				
			#Seperate by year and by color
			try:
				procYear = dateParser.parse(col[tsvDate]).year
			except:
				procYear = col[tsvDate]
			
			#Check if year & if motif exists
			if not procYear in yearMotifs.keys():
				yearMotifs[procYear] = {}
			if not motif in yearMotifs[procYear].keys():
				yearMotifs[procYear][motif] = 1
			else:
				yearMotifs[procYear][motif] += 1
	
	#Sort top level dict
			
	#Pretty print to an excel file for distribution
	wb = Workbook()
	wsOrig = wb.active
	wsOrig.title = "Motif Frequency Analysis"
	
	#Add in conditional formatting for highlights
	setProcessedConditionalFormatting(wsOrig)
	
	for key, year in yearMotifs.items():
		wsOrig.append([])
		wsOrig.append([key])
		for motif,value in year.items():
			wsOrig.append([motif, value])
			
	print("Writing to file")
	wb.save(inputfile + ".freqanalysis.xlsx")
	
if __name__ == "__main__": main()