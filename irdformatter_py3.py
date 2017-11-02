#! /usr/bin/env python

#for Python3
#DEPENDS ON OPENPYXL!
import sys
import getopt
import textwrap

import dateutil
import dateutil.parser as dateParser

import openpyxl
from openpyxl import Workbook
from openpyxl import formatting
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import (
	get_column_letter
)

#HELP DOCUMENATION
def displayHelp():
		print("""
@@ snake @@  11/96
                          _,..,,,_ 
                     '``````^~"-,_`"-,_
       .-~c~-.                    `~:. ^-.     
   `~~~-.c    ;                      `:.  `-,     _.-~~^^~:.
         `.   ;      _,--~~~~-._       `:.   ~. .~          `.
          .` ;'   .:`           `:       `:. I  `    _.:-,.    `.
        .' .:   :'    _.-~^~-.    `.       `..'   .:      `.    '
       :  .' _:'   .-'        `.    :.     .:   .'`.        :    ;
  jgs  :  `-'   .:'             `.    `^~~^`   .:.  `.      ;    ;
        `-.__,-~                  ~-.        ,' ':    '.__.`    :'
                                     ~--..--'     ':.         .:'
                                                     ':..___.:'
from: http://chris.com/ascii/index.php?art=animals/reptiles/snakes
		
USAGE: python irdformatter.py -i inputfile [-n]
		
-i Tells the location to a .TSV file downloaded from IRD for processing. This file needs to contain the following default columns; Sequence Accession, Strain Name, Collection Date. It also requires on optional column, Sequence. Be sure ot check it when downloading the file.
-n No Align. For the analysis page, a heuristic is used for pseudo alignment. Sequence that fall outside of the scope of the heuristic are thrown out to prevent as many problems as possible. No align can be used if the the data is preprocessed and the sequences are aligned.
""")

#Predefine all our colors in one place
clusters = {
    'turkey':['NHNNYR'],
    'purple':['KHHNYR','KHKNYS'],
    'gray':['KTHKYS','NTHNFK','NTQKFN','NTHKFN', 'KHKEYS'],
    'darkgray':['KTHNFK','KTHNSK'],
    'pink':['KHQKYS','NHQKYS','KHQKYR'],
    'gold':['KYNNNK'],
    'red':['NYNNYK','NYNNHK','NHNNYK','NYHNYK'],
    'red-like':['NYSNYK'],
    'darkpink':['KHKNYR'],
    'cyan':['NNNDYR', 'NHSNYR', 'NHNNYR', 'NHNDYR'],
    'brown':['NYHGHE'],
    'green':['KYKNYE'],
    'lightgreen':['KYNNYK'],
    'orange':['KHKNYE'],
    'blue':['KHNNYK'],
    'lightblue':['KHKEYS','KHHNYR'],
    'peach':['KYHNNK','KHHNNK']
}

conditionalFill = {
    'turkey': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
    'purple': PatternFill(start_color='800080', end_color='800080', fill_type='solid'),
    'gray': PatternFill(start_color='cccccc', end_color='cccccc', fill_type='solid'),
    'darkgray': PatternFill(start_color='888888', end_color='888888', fill_type='solid'),
    'pink': PatternFill(start_color='ffc0cb', end_color='ffc0cb', fill_type='solid'),
    'gold': PatternFill(start_color='f9d624', end_color='f9d624', fill_type='solid'),
    'red': PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'),
    'red-like': PatternFill(start_color='990000', end_color='990000', fill_type='solid'),
    'darkpink': PatternFill(start_color='880088', end_color='880088', fill_type='solid'),
    'cyan':PatternFill(start_color='00ffff', end_color='00ffff', fill_type='solid'),
    'brown':PatternFill(start_color='f4a460', end_color='f4a460', fill_type='solid'),
    'green':PatternFill(start_color='008800', end_color='008800', fill_type='solid'),
    'lightgreen':PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
    'orange': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
    'blue': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),
    'lightblue': PatternFill(start_color='6495ed', end_color='6495ed', fill_type='solid'),
    'peach': PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid'),
}

#Condtional Formattings for processed page
def setProcessedConditionalFormatting(wsProc):

	for key,list in clusters.items():
		for clusterMotif in list:
				dxf = DifferentialStyle(fill=conditionalFill[key])
				rule = Rule(type="containsText", operator="containsText", text=clusterMotif, dxf=dxf) #blue
				rule.formula = ['NOT(ISERROR(SEARCH("' + clusterMotif + '",D1)))']
				wsProc.conditional_formatting.add('D:D', rule)
	
#Convert sequence into a color
def assignColor(motif):
	antigenicColor = "unknown"
	
	for key,list in clusters.items():
		for clusterMotif in list:
			if(motif == clusterMotif):
				antigenicColor = key;
				return antigenicColor;

	return antigenicColor

#Variability Analysis  wsAnal:Spreadsheet   maxLen:Length of longest sequence   index: number of sequences
def variabilityAnalysis(wsAnal, maxLen, offset = 0, writePos = True):
	index = wsAnal._current_row   
	
	#Count each AA
	wsAnal.append([])
	for i in range(65, 91):
		formulas = []
		for j in range(0, maxLen):
			letter = get_column_letter(j+2)
			formulas.append("=COUNTIF(" + letter + "$1:" + letter + "$" + str(index) + ",$A" + str(index + (i - 63)) + ")")
		wsAnal.append([chr(i)] + formulas)

	#Add in a relative position count
	formulas = [''] * 17
	wsAnal.append([])
	if(writePos == True):
		for i in range(1, maxLen):
			formulas.append(str(i))
		wsAnal.append(formulas)
	else:
		wsAnal.append([])

	#Find Max, relative position calculations 2 blanks and 26 letters
	formulas = []
	for j in range(0, maxLen):
		letter = get_column_letter(j+2)
		formulas.append("=MAX(" + letter + str(index + 2) + ":" + letter + str(index + 1 + 26) + ")" )
	wsAnal.append(["MAX"] + formulas)

	#Calc Variability Calculating by index is not entirely correct, should find the max of the row. Will have none = 1
	wsAnal.append([])
	formulas = []
	buffer = [''] * offset
	for j in range(0, maxLen):
		letter = get_column_letter(j + 2 + offset)
		formulas.append("=" + letter +  str(index + 4 + 26) + "/" + str(index))
	wsAnal.append(buffer + ["Variability"] + formulas)

	#Add in conditional formatting
	rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFFF0000', end_type='percentile', end_value=100, end_color='FFFFFFFF')
	wsAnal.conditional_formatting.add("B" + str(index + 6 + 26) + ":" + get_column_letter(maxLen+1) + str(index + 6 + 26), rule)

def main():
	#Get arguments
	argv = sys.argv[1:]
	inputfile = ''		#Input .tsv as STRING
	vAnalysis = False	#whether to perform variability anlysis as BOOLEAN
	heuristicAlign = True
	cAnalysis = False
	try:
		opts, args = getopt.getopt(argv,"i:vnch",["input=","vanalysis","noalign","coloranalysis","help"])
	except getopt.GetoptError:
		displayHelp()
		sys.exit(2)
	for opt, arg in opts:
		if opt in ('-h',"help"):
			displayHelp()
			sys.exit()
		elif opt in ("-i", "--ifile"):
			inputfile = arg
		elif opt in ("-n", "--noalign"):
			heuristicAlign = False
		elif opt in ("-v", "--vanalysis"):
			vAnalysis = True
		elif opt in ("-c", "--coloranalysis"):
			cAnalysis = True

	#Create a new 3 page workbook
	wb = Workbook()

	#Create our pages
	wsOrig = wb.active
	wsOrig.title = "OriginalData"
	wsProc = wb.create_sheet(title="Processed")
	wsAnal = wb.create_sheet(title="VariabilityAnalysis")

	#Create coloranalysis workbook
	if(cAnalysis == True):
		cwb = Workbook()
		cwbPages = {}
		
	#Add in conditional formatting
	setProcessedConditionalFormatting(wsProc)

	#maxLen mechanism
	maxLen = 0

	#Load input for line by line reading
	with open(inputfile) as f:
		for index, line in enumerate(f):
			col = []
			col = line.split("\t")
			
			#Get Indices from first line
			if(index ==0):
				tsvAccession = col.index("Sequence Accession")
				tsvDate = col.index("Collection Date")
				tsvName = col.index("Strain Name")
				tsvSeq = col.index("Sequence")
				
				#Write headers on all pages
				wsOrig.append(col);
				wsProc.append(['Isolate Name', 'Collection Date', 'Year', 'Antigenic Motif', '145', '155', '156', '158', '159', '189', 'Clade', 'Duplicate', 'Color'])
			
				#Next iteration
				continue

			#ALIGN -- Use heuristics to 'align' sequences on the fly. Based on calculating offset by identifying the most stable part(s) of the sequence, then shifts left. Will fail on indel
			if(heuristicAlign == True):
				if(col[tsvSeq][0] != "M"):	#96% start properly
					seqShift = col[tsvSeq].upper().find('FSRLNWL')			#Find the relative position of 148(+16)-154(+16) and shift right with spaces
					if(seqShift < (148+16)):
						charFill = ' '*((148+16) - seqShift - 1)
						col[tsvSeq] = charFill + col[tsvSeq]
					elif(seqShift == -1):									 #can either try another relative position or scrap the sequence. I am lazy this time, so I am down with scrapping. Need to insert blanks to keep sheet functioning on the index
						wsOrig.append([])
						wsProc.append([])
						continue

			#Copy information to first sheet, in the less then pythonic way (could use the csv module...)
			wsOrig.append(col);

			#Process the data so it matches Jered's format, to some extent
			procIsolateName = col[tsvName] + "|" + col[tsvAccession]
			try:
				procYear = dateParser.parse(col[tsvDate]).year
			except:
				procYear = col[tsvDate]
		
			procMotif = "=CONCATENATE(E"+ str(index + 1) +",F"+ str(index + 1) +",G"+ str(index + 1) +",H"+ str(index + 1) +",I"+ str(index + 1) +",J"+ str(index + 1) +")"
			proc145 = "=MID(OriginalData!" + get_column_letter(tsvSeq + 1) + str(index + 1) + ",145+16,1)"  #Python indices start at 0, excel at 1
			proc155 = "=MID(OriginalData!" + get_column_letter(tsvSeq + 1) + str(index + 1) + ",155+16,1)"
			proc156 = "=MID(OriginalData!" + get_column_letter(tsvSeq + 1) + str(index + 1) + ",156+16,1)"
			proc158 = "=MID(OriginalData!" + get_column_letter(tsvSeq + 1) + str(index + 1) + ",158+16,1)"
			proc159 = "=MID(OriginalData!" + get_column_letter(tsvSeq + 1) + str(index + 1) + ",159+16,1)"
			proc189 = "=MID(OriginalData!" + get_column_letter(tsvSeq + 1) + str(index + 1) + ",189+16,1)"
		
			#Generate motif to figure out color
			try:
				motif = col[tsvSeq][145+16-1] + col[tsvSeq][155+16-1] + col[tsvSeq][156+16-1] + col[tsvSeq][158+16-1] + col[tsvSeq][159+16-1] + col[tsvSeq][189+16-1]
			except:
				motif = ''
			
			#GET COLOR
			antigenicColor = assignColor(motif)
			
			wsProc.append([procIsolateName,col[tsvDate], procYear, procMotif, proc145, proc155, proc156, proc158, proc159, proc189, '', '', antigenicColor])

			#Max Length Mechanism
			if(len(col[tsvSeq]) > maxLen):
				maxLen = len(col[tsvSeq]) - 1

			#Copy sequences to 3rd page
			wsAnal.append([''] + list(col[tsvSeq]))
			
			#Add to color analysis
			if(cAnalysis == True):
				#Check if page exists, if not create
				if(antigenicColor not in cwbPages.keys()):
					cwbPages[antigenicColor] = cwb.create_sheet(title = antigenicColor)
				cwbPages[antigenicColor].append([procIsolateName, procYear, antigenicColor,] + list(col[tsvSeq][16:230]))
			 
	#Variability analysis
	variabilityAnalysis(wsAnal,maxLen)
	
	# Save the file
	print("Writing to file")
	wb.save(inputfile + ".xlsx")

	'''
	#Copy and paste the results to one page for easy viewing, this is not the clean way
	cwbMain = cwb.active
	cwbMain.title = "Overview"
	for i, sheet in enumerate(cwbPages):
		variabilityAnalysis(cwbPages[sheet], 214, 2)
		for j in range(1, 214):
			cwbMain.cell(row = i + 1, column = j).value = cwbPages[sheet].cell(row = cwbPages[sheet]._current_row - 1, column = j).value
	'''		

	

	#Save coloranalysis workbook
	if(cAnalysis == True):
		cwb.save(inputfile + "color.xlsx")

if __name__ == "__main__": main()

